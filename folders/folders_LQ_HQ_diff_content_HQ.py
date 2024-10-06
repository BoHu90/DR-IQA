"""
    folders_LQ_HQ_NAR 读取LQ HQ NAR 图像
    对于合成失真数据集，返回LQ、HQ图片 以及 NAR（DIV2K)图片 和LQ的标签
    对于真实失真数据集，则没有HQ图片
"""

import torch
import torch.utils.data as data
import torchvision
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
import random
from openpyxl import load_workbook
from PIL import ImageFile

from Utils.ImageSelector import ImageSelector

ImageFile.LOAD_TRUNCATED_IMAGES = True  # 防止出现OS读取图像报错
import pandas as pd

from Utils.addDistortions import apply_classical_distortions, apply_random_distortions, apply_classical_distortions1


class Kadid10kFolder(data.Dataset):
    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        LQ_paths = []
        HQ_paths = []
        mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                LQ_paths.append(row['dist_img'])
                HQ_paths.append(row['ref_img'])
                mos = np.array(float(row['dmos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, 'imgs', LQ_paths[item]),
                               os.path.join(root, 'imgs', HQ_paths[item]), mos_all[item]))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, ZLQ_content, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO Kadid HQ\ZLQ 的获取
        ZLQ_content = apply_classical_distortions(LQ)  # 固定10种
        # HQ = pil_loader(HQ_path)  # 参考图像
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]  # 随机选取
        # HQ_path = self.selector.select_similar_image(LQ)  # 最相似
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        # 切patch
        LQ_patches, HQ_patches, ZLQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            # LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)

            LQ_patch = self.ZLQ_transform(LQ)
            HQ_patch = self.ZLQ_transform(HQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ_content)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class LIVEFolder(data.Dataset):

    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num
        self.root = root

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos.mat'))
        labels = dmos['dmos'].astype(np.float32)

        orgs = dmos['orgs']
        refpaths_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refpaths_all = refpaths_all['refnames_all']

        sample = []
        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = imgpath[item]
                    HQ_path = os.path.join(root, 'refimgs', refpaths_all[0][item][0])
                    label = labels[0][item]
                    sample.append((LQ_path, HQ_path, label))
                # print(self.imgpath[item])

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        # 选择器
        # self.selector = ImageSelector(self.ref_paths)

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO LIVE HQ\ZLQ 的获取
        # HQ = pil_loader(HQ_path)  # 参考图像
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]  # 随机选取
        # HQ_path = self.selector.select_similar_image(LQ)
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        ZLQ = apply_classical_distortions(LQ)  # 固定10种

        # LIVE 随机10个patches
        LQ_patches, HQ_patches, ZLQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            # LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)

            LQ_patch = self.ZLQ_transform(LQ)
            HQ_patch = self.ZLQ_transform(HQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)
        return LQ_patches, HQ_patches, ZLQ_patches, target

        # LQ = self.ZLQ_transform(LQ)
        # HQ = self.ZLQ_transform(HQ)
        # ZLQ = self.ZLQ_transform(ZLQ)
        # return LQ, HQ, ZLQ, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


class CSIQFolder(data.Dataset):

    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath, '.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        LQ_pathes = []
        target = []
        refpaths_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            LQ_pathes.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refpaths_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refpaths_all = np.array(refpaths_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'dst_imgs_all', LQ_pathes[item])
                    HQ_path = os.path.join(root, 'src_imgs', refpaths_all[item])
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO CSIQ HQ\ZLQ 的获取
        # HQ = pil_loader(HQ_path)  # 参考图像
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        ZLQ = apply_classical_distortions(LQ)  # 固定10
        LQ_patches, HQ_patches, ZLQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            # LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)

            LQ_patch = self.ZLQ_transform(LQ)
            HQ_patch = self.ZLQ_transform(HQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class TID2013Folder(data.Dataset):

    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        refpath = os.path.join(root, 'reference_images')
        refname = self._getTIDFileName(refpath, '.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        LQ_pathes = []
        target = []
        refpaths_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            LQ_pathes.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refpaths_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refpaths_all = np.array(refpaths_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refpaths_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'distorted_images', LQ_pathes[item])
                    refHQ_name = 'I' + LQ_pathes[item].split("_")[0][1:] + '.BMP'
                    HQ_path = os.path.join(refpath, refHQ_name)
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def _getTIDFileName(self, path, suffix):
        filename = []
        f_list = os.listdir(path)
        for i in f_list:
            if suffix.find(os.path.splitext(i)[1]) != -1:
                filename.append(i[1:3])
        return filename

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]

        LQ = pil_loader(LQ_path)
        # HQ = pil_loader(HQ_path)  # 参考图像
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        # TODO TID HQ\ZLQ 的获取
        ZLQ = apply_classical_distortions(LQ)  # 固定10

        LQ_patches, HQ_patches, ZLQ_patches = [], [], []
        for _ in range(self.self_patch_num):  # 打patch
            # LQ_patch, HQ_patch = getPairRandomPatch(LQ, HQ, crop_size=self.patch_size)

            LQ_patch = self.ZLQ_transform(LQ)
            HQ_patch = self.ZLQ_transform(HQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class LIVEChallengeFolder(data.Dataset):
    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        LQ_pathes = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        LQ_pathes = LQ_pathes['AllImages_release']
        LQ_pathes = LQ_pathes[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, 'Images', LQ_pathes[item][0][0]), labels[item]))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        # TODO LIVEC HQ\ZLQ 的获取
        LQ = pil_loader(LQ_path)

        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)

        ZLQ = apply_classical_distortions(LQ)
        LQ_patches, ZLQ_patches, HQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.ZLQ_transform(LQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)
            HQ_patch = self.ZLQ_transform(HQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class BIDChallengeFolder(data.Dataset):
    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        LQ_pathes = []
        labels = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for _ in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "ImageDatabase/DatabaseImage%04d.JPG" % (img_num)
            LQ_pathes.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            labels.append(mos)
            if count == 587:
                break

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, LQ_pathes[item]), labels[item]))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO BID HQ\ZLQ 的获取
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        ZLQ = apply_classical_distortions(LQ)  # 固定10
        LQ_patches, ZLQ_patches, HQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.ZLQ_transform(LQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)
            HQ_patch = self.ZLQ_transform(HQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq_10kFolder(data.Dataset):
    def __init__(self, root, ref_root, index, transform, ZLQ_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, '1024x768', imgname[item]), mos_all[item]))

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, _, ZLQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO koniq HQ\ZLQ 的获取
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        ZLQ = apply_classical_distortions(LQ)  # 固定10
        LQ_patches, ZLQ_patches, HQ_patches = [], [], []
        for _ in range(self.self_patch_num):
            LQ_patch = self.ZLQ_transform(LQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)
            HQ_patch = self.ZLQ_transform(HQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)

        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


class SPAQFolder(data.Dataset):
    def __init__(self, root, index, transform, HQ_diff_content_transform, patch_num,
                 patch_size=224, self_patch_num=10):
        super(SPAQFolder, self).__init__()

        self.self_patch_num = self_patch_num
        self.data_path = root
        xlsx_file = os.path.join(self.data_path, "MOS and Image attribute scores.xlsx")
        read = pd.read_excel(xlsx_file)
        imgname = read["Image name"].values.tolist()
        mos_all = read["MOS"].values.tolist()
        for i in range(len(mos_all)):
            mos_all[i] = np.array(mos_all[i]).astype(np.float32)
        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append(
                    (
                        os.path.join(
                            self.data_path,
                            "512x384",
                            imgname[item],
                        ),
                        mos_all[item],
                    )
                )

        self.samples = sample
        self.transform = transform
        self.HQ_diff_content_transform = HQ_diff_content_transform

    def _load_image(self, path):
        try:
            im = Image.open(path).convert("RGB")
        except: # 出现错误，则返回随机im
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path)
        # TODO SPAQ HQ_Diff HQ\ZLQ 的获取
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        HQ_diff_content = apply_classical_distortions(LQ)  # 固定10
        LQ_patches, HQ_diff_content_patches = [], []
        HQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.HQ_diff_content_transform(LQ)
            HQ_diff_content_patch = self.HQ_diff_content_transform(HQ_diff_content)
            HQ_patch = self.HQ_diff_content_transform(HQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            HQ_diff_content_patches.append(HQ_diff_content_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
        # [self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_diff_content_patches = torch.cat(HQ_diff_content_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)

        return LQ_patches, HQ_patches, HQ_diff_content_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


#FLive数据集
class FLiveFolder(data.Dataset):
    def __init__(self, root, ref_root, index, transform, patch_num, ZLQ_transform,
                 patch_size=224, self_patch_num=10):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num

        path = os.path.join(root, 'labels_image.csv')

        df_ims = pd.read_csv(path)
        # 从csv文件中获取图像路径
        image_path = df_ims['name']
        image_path = image_path.values.tolist()
        # 从csv文件中获取mos
        mos = df_ims['mos']
        mos = mos.values.tolist()

        print(f"Loaded image paths: {len(image_path)}, and index is {len(index)}")  # 添加调试信息

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append(((os.path.join(root, image_path[item])), mos[item]))
        # # 打印sample一个元素的类型和值
        # print(f"sample type: {type(sample[0])}, value: {sample[0]}")

        # self.ref_paths = []
        # for HQ_diff_content_img_path in os.listdir(ref_root):
        #     if HQ_diff_content_img_path[-3:] == 'png' or HQ_diff_content_img_path[
        #                                                  -3:] == 'jpg' or HQ_diff_content_img_path[-3:] == 'bmp':
        #         self.ref_paths.append(os.path.join(ref_root, HQ_diff_content_img_path))

        self.samples = sample
        print('--FLive：', len(self.samples))
        self.transform = transform
        self.ZLQ_transform = ZLQ_transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]  # 读取路径和标签
        # # 打印path和target的类型和值
        # print(f"path type: {type(path)}, value: {path}")
        # print(f"target type: {type(target)}, value: {target}")

        LQ = pil_loader(path)  # 读取图像
        # TODO FLIVE HQ\ZLQ 的获取
        # 非对齐参考图像
        # HQ_path = self.ref_paths[random.randint(0, len(self.ref_paths) - 1)]
        HQ_path = '/data/dataset/DIV2K/train_HR_Sample/0535.png'
        HQ = pil_loader(HQ_path)
        ZLQ = apply_classical_distortions(LQ)  # 得到再失真图

        LQ_patches, ZLQ_patches, HQ_patches = [], [], []
        for _ in range(self.self_patch_num):  # 随机裁块
            LQ_patch = self.ZLQ_transform(LQ)
            ZLQ_patch = self.ZLQ_transform(ZLQ)
            HQ_patch = self.ZLQ_transform(HQ)

            LQ_patches.append(LQ_patch.unsqueeze(0))
            ZLQ_patches.append(ZLQ_patch.unsqueeze(0))
            HQ_patches.append(HQ_patch.unsqueeze(0))
        LQ_patches = torch.cat(LQ_patches, 0)
        HQ_patches = torch.cat(HQ_patches, 0)
        ZLQ_patches = torch.cat(ZLQ_patches, 0)
        return LQ_patches, HQ_patches, ZLQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


# crop操作
def getPairRandomPatch(img1, img2, crop_size=512):
    (iw, ih) = img1.size
    # print(ih,iw)

    ip = int(crop_size)

    ix = random.randrange(0, iw - ip + 1)
    iy = random.randrange(0, ih - ip + 1)

    img1_patch = img1.crop((ix, iy, ix + ip, iy + ip))  # 左上右下
    img2_patch = img2.crop((ix, iy, ix + ip, iy + ip))  # 左上右下

    return img1_patch, img2_patch


def getPairAugment(img1, img2, hflip=True, vflip=False, rot=False):
    hflip = hflip and random.random() < 0.5
    vflip = vflip and random.random() < 0.5
    rot180 = rot and random.random() < 0.5

    if hflip:
        img1 = img1.transpose(Image.FLIP_TOP_BOTTOM)
        img2 = img2.transpose(Image.FLIP_TOP_BOTTOM)
    if vflip:
        img1 = img1.transpose(Image.FLIP_LEFT_RIGHT)
        img2 = img2.transpose(Image.FLIP_LEFT_RIGHT)
    if rot180:
        img1 = img1.transpose(Image.ROTATE_180)
        img2 = img2.transpose(Image.ROTATE_180)

    return img1, img2


def getSelfPatch(img, patch_size, patch_num, is_random=True):
    (iw, ih) = img.size
    patches = []
    for i in range(patch_num):
        if is_random:
            ix = random.randrange(0, iw - patch_size + 1)
            iy = random.randrange(0, ih - patch_size + 1)
        else:
            ix, iy = (iw - patch_size + 1) // 2, (ih - patch_size + 1) // 2

        # patch = img[iy:iy + lr_size, ix:ix + lr_size, :]#上下左右
        patch = img.crop((ix, iy, ix + patch_size, iy + patch_size))  # 左上右下
        patches.append(patch)

    return patches


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')


